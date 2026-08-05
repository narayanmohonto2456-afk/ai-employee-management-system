from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


def generate_attendance_excel(queryset):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Attendance Report"

    headers = [
        "Employee ID",
        "Employee Name",
        "Department",
        "Date",
        "Check In",
        "Check Out",
        "Working Hours",
        "Status",
    ]

    for column, header in enumerate(headers, start=1):

        cell = worksheet.cell(row=1, column=column)

        cell.value = header

        cell.font = Font(bold=True)

    row = 2

    for attendance in queryset:

        worksheet.cell(row=row, column=1).value = attendance.employee.employee_id

        worksheet.cell(row=row, column=2).value = (
            attendance.employee.user.get_full_name()
        )

        worksheet.cell(row=row, column=3).value = (
            attendance.employee.department.department_name
        )

        worksheet.cell(row=row, column=4).value = str(attendance.date)

        worksheet.cell(row=row, column=5).value = (
            str(attendance.check_in) if attendance.check_in else "-"
        )

        worksheet.cell(row=row, column=6).value = (
            str(attendance.check_out) if attendance.check_out else "-"
        )

        worksheet.cell(row=row, column=7).value = (
            attendance.working_hours
        )

        worksheet.cell(row=row, column=8).value = attendance.status

        row += 1

    return workbook

def generate_attendance_pdf(response, queryset):

    document = SimpleDocTemplate(
        response,
        pagesize=A4,
    )

    data = [[
        "Employee ID",
        "Employee",
        "Department",
        "Date",
        "Status",
    ]]

    for attendance in queryset:

        data.append([
            attendance.employee.employee_id,
            attendance.employee.user.get_full_name(),
            attendance.employee.department.department_name,
            str(attendance.date),
            attendance.status,
        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        ])

    )

    document.build([table])